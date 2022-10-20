from datetime import datetime, timedelta
from openpyxl import load_workbook

#設定檔案
NA='NA.xlsx'
NB='NB.xlsx'
NE='NE.xlsx'
NP='NP.xlsx'
result='o.xlsx'

today=datetime.strftime(datetime.today(),'%Y/%m/%d')
yesterday=datetime.strftime(datetime.today()+timedelta(days=-1),'%Y/%m/%d')
resultM=[]
resultP=[]

NA=load_workbook(filename=NA).active
NA_target=[]
NA_M=[]
NA_P=[]
for row in NA['B1':'B1000']:
    for cell in row:
        if isinstance(cell.value, datetime):
            if datetime.strftime(cell.value,'%Y/%m/%d')==today:
                NA_target+=[cell.row]
for i in range(len(NA_target)-1):
    if NA_target[i+1]-NA_target[i]!=1:
        NA_M=NA_target[:i+1]
        NA_P=NA_target[i+1:]
for i in NA_M:
    resultM+=[(NA['H'+str(i)].value,NA['I'+str(i)].value,'NA')]
for i in NA_P:
    resultP+=[(NA['H'+str(i)].value,NA['I'+str(i)].value,'NA')]


NB=load_workbook(filename=NB).active
NB_target=[]
NB_M=[]
NB_P=[]
for row in NB['B1':'B300']:
    for cell in row:
            if isinstance(cell.value, datetime):
                if datetime.strftime(cell.value,'%Y/%m/%d')==today:
                    NB_target+=[cell.row]
            elif isinstance(cell.value, str):
                if cell.value==today:
                    NB_target+=[cell.row]
for i in range(len(NB_target)-1):
    if NB_target[i+1]-NB_target[i]!=1:
        NB_M=NB_target[:i+1]
        NB_P=NB_target[i+1:]
for i in NB_M:
    resultM+=[(NB['H'+str(i)].value,NB['I'+str(i)].value,'NB')]
for i in NB_P:
    resultP+=[(NB['H'+str(i)].value,NB['I'+str(i)].value,'NB')]


NE=load_workbook(filename=NE).active
NE_target=[]
for row in NE['B1':'B100']:
    for cell in row:
        if isinstance(cell.value, datetime):
            if datetime.strftime(cell.value,'%Y/%m/%d')==today:
                NE_target+=[cell.row]
for i in NE_target:
    resultM+=[(NE['H'+str(i)].value,NE['I'+str(i)].value,'NE')]


NP=load_workbook(filename=NP).active
NP_target=[]
NP_M=[]
NP_P=[]
for row in NP['B1':'B500']:
    for cell in row:
        if isinstance(cell.value, datetime):
            if datetime.strftime(cell.value,'%Y/%m/%d')==today:
                NP_target+=[cell.row]
for i in range(len(NP_target)-1):
    if NP_target[i+1]-NP_target[i]!=1:
        NP_M=NP_target[:i+1]
        NP_P=NP_target[i+1:]
for i in NP_M:
    resultM+=[(NP['H'+str(i)].value,NP['I'+str(i)].value,'NP')]
for i in NP_P:
    resultP+=[(NP['H'+str(i)].value,NP['I'+str(i)].value,'NP')]


#寫入檔案
wb=load_workbook(filename=result)
result=wb.active
result_target=[]
for row in result['B1':'B500']:
    for cell in row:
        if cell.value!=None:
            result_target+=[cell.row]
R=max(result_target)+2

#寫入標題
lst=['序號','申報處方日期','開立處方日期','開立醫療機構名稱','藥物類別',\
'民眾姓名','民眾身分證字號','庫台','用量','餘量']
for i in range(len(lst)):
    result[chr(65+i)+str(R)]=lst[i]

#寫入M
for i in range(1,len(resultM)+1):
    result['A'+str(i+R)]=i
    result['B'+str(i+R)]=today
    result['C'+str(i+R)]=yesterday
    result['D'+str(i+R)]='高雄長庚紀念醫院'
    result['E'+str(i+R)]='Molnupiravir'
    result['F'+str(i+R)]=resultM[i-1][0]
    result['G'+str(i+R)]=resultM[i-1][1]
    result['H'+str(i+R)]=resultM[i-1][2]
result['I'+str(i+R)]=len(resultM)

#寫入P
R+=len(resultM)
for i in range(1,len(resultP)+1):
    result['A'+str(i+R)]=i
    result['B'+str(i+R)]=today
    result['C'+str(i+R)]=yesterday
    result['D'+str(i+R)]='高雄長庚紀念醫院'
    result['E'+str(i+R)]='Paxlovid'
    result['F'+str(i+R)]=resultP[i-1][0]
    result['G'+str(i+R)]=resultP[i-1][1]
    result['H'+str(i+R)]=resultP[i-1][2]
result['I'+str(i+R)]=len(resultP)

wb.save('result.xlsx')