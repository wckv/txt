from os import listdir
from tkinter import Entry, Radiobutton, messagebox, Tk, StringVar, Label, Button
#from ttkbootstrap import Style
from datetime import datetime,timedelta
from openpyxl import Workbook, load_workbook

def GButton_28_command():
    try:
        filevar=[varNA.get(),varNB.get(),varNE.get(),varNP.get()]
        filename=[GLineEdit_496.get(),GLineEdit_930.get(),GLineEdit_862.get(),GLineEdit_895.get()]
        todaytitle=datetime.strftime(datetime.today(),'%m%d')
        tabtitle=['NA'+todaytitle,'NB'+todaytitle,'NE'+todaytitle,'NP'+todaytitle]
        if '1' not in filevar:
            messagebox.showwarning(title='四個不存在', message='都沒有使用就不用執行了吧？\n\nver. 1.0')
        elif set([filename[x] for x in range(4) if filevar[x]=='1'])<= set(listdir()):
            result=Workbook()
            rs=result.active
            rs.title='工作區'
            for i in range(4):
                if filevar[i]=='1':
                    #建立新分頁
                    rsi=result.create_sheet(tabtitle[i])
                    open=load_workbook(filename=filename[i]).active
                    for column in range(20):
                        for row in range(1,open.max_row+1):
                            rsi[chr(65+column)+str(row)].value=open[chr(65+column)+str(row)].value

            #將內容複製至工作分頁
            today=datetime.strftime(datetime.today(),'%Y/%m/%d')
            yesterday=datetime.strftime(datetime.today()+timedelta(days=-1),'%Y/%m/%d')
            resultM=[]
            resultP=[]

            if filevar[0]=='1':
                NA=result['NA'+todaytitle]
                NA_target=[]
                NA_M=[]
                NA_P=[]
                for row in NA['B1':'B'+str(NA.max_row)]:
                    for cell in row:
                        if isinstance(cell.value, datetime):
                            if datetime.strftime(cell.value,'%Y/%m/%d')==today:
                                NA_target+=[cell.row]
                for i in range(len(NA_target)-1):
                    if NA_target[i+1]-NA_target[i]!=1:
                        NA_M=NA_target[:i+1]
                        NA_P=NA_target[i+1:]
                for i in NA_M:
                    resultM+=[(NA['H'+str(i)].value,NA['I'+str(i)].value,'NA')]
                for i in NA_P:
                    resultP+=[(NA['H'+str(i)].value,NA['I'+str(i)].value,'NA')]

            if filevar[1]=='1':
                NB=result['NB'+todaytitle]
                NB_target=[]
                NB_M=[]
                NB_P=[]
                for row in NB['B1':'B'+str(NB.max_row)]:
                    for cell in row:
                            if isinstance(cell.value, datetime):
                                if datetime.strftime(cell.value,'%Y/%m/%d')==today:
                                    NB_target+=[cell.row]
                            elif isinstance(cell.value, str):
                                if cell.value==today:
                                    NB_target+=[cell.row]
                for i in range(len(NB_target)-1):
                    if NB_target[i+1]-NB_target[i]!=1:
                        NB_M=NB_target[:i+1]
                        NB_P=NB_target[i+1:]
                for i in NB_M:
                    resultM+=[(NB['H'+str(i)].value,NB['I'+str(i)].value,'NB')]
                for i in NB_P:
                    resultP+=[(NB['H'+str(i)].value,NB['I'+str(i)].value,'NB')]

            if filevar[2]=='1':
                NE=result['NE'+todaytitle]
                NE_target=[]
                for row in NE['B1':'B'+str(NE.max_row)]:
                    for cell in row:
                        if isinstance(cell.value, datetime):
                            if datetime.strftime(cell.value,'%Y/%m/%d')==today:
                                NE_target+=[cell.row]
                for i in NE_target:
                    resultM+=[(NE['H'+str(i)].value,NE['I'+str(i)].value,'NE')]

            if filevar[3]=='1':
                NP=result['NP'+todaytitle]
                NP_target=[]
                NP_M=[]
                NP_P=[]
                for row in NP['B1':'B'+str(NP.max_row)]:
                    for cell in row:
                        if isinstance(cell.value, datetime):
                            if datetime.strftime(cell.value,'%Y/%m/%d')==today:
                                NP_target+=[cell.row]
                for i in range(len(NP_target)-1):
                    if NP_target[i+1]-NP_target[i]!=1:
                        NP_M=NP_target[:i+1]
                        NP_P=NP_target[i+1:]
                for i in NP_M:
                    resultM+=[(NP['H'+str(i)].value,NP['I'+str(i)].value,'NP')]
                for i in NP_P:
                    resultP+=[(NP['H'+str(i)].value,NP['I'+str(i)].value,'NP')]


            #寫入標題
            lst=['序號','申報處方日期','開立處方日期','開立醫療機構名稱','藥物類別',\
            '民眾姓名','民眾身分證字號','庫台','用量','餘量']
            for i in range(len(lst)):
                rs[chr(65+i)+'1']=lst[i]

            #寫入M
            for i in range(2,len(resultM)+2):
                rs['A'+str(i)]=i-1
                rs['B'+str(i)]=today
                rs['C'+str(i)]=yesterday
                rs['D'+str(i)]='高雄長庚紀念醫院'
                rs['E'+str(i)]='Molnupiravir'
                rs['F'+str(i)]=resultM[i-2][0]
                rs['G'+str(i)]=resultM[i-2][1]
                rs['H'+str(i)]=resultM[i-2][2]
            rs['I'+str(len(resultM)+1)]=len(resultM)

            #寫入P
            for i in range(1,len(resultP)+1):
                rs['A'+str(i+len(resultM)+1)]=i
                rs['B'+str(i+len(resultM)+1)]=today
                rs['C'+str(i+len(resultM)+1)]=yesterday
                rs['D'+str(i+len(resultM)+1)]='高雄長庚紀念醫院'
                rs['E'+str(i+len(resultM)+1)]='Paxlovid'
                rs['F'+str(i+len(resultM)+1)]=resultP[i-1][0]
                rs['G'+str(i+len(resultM)+1)]=resultP[i-1][1]
                rs['H'+str(i+len(resultM)+1)]=resultP[i-1][2]
            rs['I'+str(len(resultP)+len(resultM)+1)]=len(resultP)

            result.save('執行結果.xlsx')
            messagebox.showinfo(title='執行完畢',message='檔案儲存完成，程式即將關閉')
            root.destroy()
        else:
            messagebox.showerror(title='找不到檔案', message='請檢查檔案是否存在，或者檔名有無錯誤')
    except:
        messagebox.showerror(title='檔案錯誤', message='是不是EXCEL檔案仍開啟中?\n\n請關閉檔案再執行')

def GButton_979_command():
    root.destroy()

#GUI
root=Tk()
#style=Style(theme='sandstone')
varNA=StringVar()
varNB=StringVar()
varNE=StringVar()
varNP=StringVar()

#選擇時顏色變化
def C1():
    GLineEdit_496['bg']='#ffffff'
    GLineEdit_496['state']='normal'
def C2():
    GLineEdit_930['bg']='#ffffff'
    GLineEdit_930['state']='normal'
def C3():
    GLineEdit_862['bg']='#ffffff'
    GLineEdit_862['state']='normal'
def C4():
    GLineEdit_895['bg']='#ffffff'
    GLineEdit_895['state']='normal'
def C5():
    GLineEdit_496['bg']='#dcdcdc'
    GLineEdit_496['state']='disabled'
def C6():
    GLineEdit_930['bg']='#dcdcdc'
    GLineEdit_930['state']='disabled'
def C7():
    GLineEdit_862['bg']='#dcdcdc'
    GLineEdit_862['state']='disabled'
def C8():
    GLineEdit_895['bg']='#dcdcdc'
    GLineEdit_895['state']='disabled'

#主畫面
def GUI_root():
    #setting title
    root.title("Excel 合併轉換")
    #setting window size
    width=500
    height=400
    screenwidth = root.winfo_screenwidth()
    screenheight = root.winfo_screenheight()
    alignstr = '%dx%d+%d+%d' % (width, height, (screenwidth - width) / 2, (screenheight - height) / 2)
    root.geometry(alignstr)
    root.resizable(width=False, height=False)
GUI_root()
#按鈕
def GUI_Button():
    GButton_28=Button(root)
    GButton_28["font"] = ('微軟正黑體',14)
    GButton_28["justify"] = "center"
    GButton_28["text"] = "確認"
    GButton_28.place(x=257,y=340,width=100,height=40)
    GButton_28["command"] = GButton_28_command

    GButton_979=Button(root)
    GButton_979["font"] = ('微軟正黑體',14)
    GButton_979["justify"] = "center"
    GButton_979["text"] = "關閉"
    GButton_979.place(x=377,y=340,width=100,height=40)
    GButton_979["command"] = GButton_979_command
GUI_Button()
#四個左選項
def Left1():
    GRadio_748=Radiobutton(root)
    GRadio_748["font"] = ('微軟正黑體',12)
    GRadio_748["justify"] = "left"
    GRadio_748["text"] = "NA檔名:"
    GRadio_748['variable']=varNA
    GRadio_748['value']='1'
    GRadio_748['command']=C1
    GRadio_748.select()
    GRadio_748.place(x=20,y=110,width=90,height=30)

    GRadio_578=Radiobutton(root)
    GRadio_578["font"] = ('微軟正黑體',12)
    GRadio_578["justify"] = "left"
    GRadio_578["text"] = "NB檔名:"
    GRadio_578['variable']=varNB
    GRadio_578['value']='1'
    GRadio_578.select()
    GRadio_578['command']=C2
    GRadio_578.place(x=20,y=170,width=90,height=30)

    GRadio_746=Radiobutton(root)
    GRadio_746["font"] = ('微軟正黑體',12)
    GRadio_746["justify"] = "left"
    GRadio_746["text"] = "NE檔名:"
    GRadio_746['variable']=varNE
    GRadio_746['value']='1'
    GRadio_746.select()
    GRadio_746['command']=C3
    GRadio_746.place(x=20,y=230,width=90,height=30)

    GRadio_112=Radiobutton(root)
    GRadio_112["font"] = ('微軟正黑體',12)
    GRadio_112["justify"] = "left"
    GRadio_112["text"] = "NP檔名:"
    GRadio_112['variable']=varNP
    GRadio_112['value']='1'
    GRadio_112.select()
    GRadio_112['command']=C4
    GRadio_112.place(x=20,y=290,width=90,height=30)
Left1()
#四個左輸入區
def Left2():
    global GLineEdit_496,GLineEdit_930,GLineEdit_862,GLineEdit_895

    GLineEdit_496=Entry(root)
    GLineEdit_496["borderwidth"] = "1px"
    GLineEdit_496["font"] = ('微軟正黑體',12)
    GLineEdit_496["justify"] = "left"
    GLineEdit_496.insert(0,"NA.xlsx")
    GLineEdit_496.place(x=120,y=110,width=250,height=30)

    GLineEdit_930=Entry(root)
    GLineEdit_930["borderwidth"] = "1px"
    GLineEdit_930["font"] = ('微軟正黑體',12)
    GLineEdit_930["justify"] = "left"
    GLineEdit_930.insert(0,"NB.xlsx")
    GLineEdit_930.place(x=120,y=170,width=250,height=30)

    GLineEdit_862=Entry(root)
    GLineEdit_862["borderwidth"] = "1px"
    GLineEdit_862["font"] = ('微軟正黑體',12)
    GLineEdit_862["justify"] = "left"
    GLineEdit_862.insert(0,"NE.xlsx")
    GLineEdit_862.place(x=120,y=230,width=250,height=30)

    GLineEdit_895=Entry(root)
    GLineEdit_895["borderwidth"] = "1px"
    GLineEdit_895["font"] = ('微軟正黑體',12)
    GLineEdit_895["justify"] = "left"
    GLineEdit_895.insert(0,"NP.xlsx")
    GLineEdit_895.place(x=120,y=290,width=250,height=30)
Left2()
#四個右選項
def Right():
    GRadio_397=Radiobutton(root)
    GRadio_397["font"]=('微軟正黑體',12)
    GRadio_397["justify"] = "left"
    GRadio_397["text"] = "NA無使用"
    GRadio_397['variable']=varNA
    GRadio_397['value']='0'
    GRadio_397['command']=C5
    GRadio_397.place(x=380,y=110,width=100,height=30)

    GRadio_980=Radiobutton(root)
    GRadio_980["font"] = ('微軟正黑體',12)
    GRadio_980["justify"] = "left"
    GRadio_980["text"] = "NB無使用"
    GRadio_980['variable']=varNB
    GRadio_980['value']='0'
    GRadio_980['command']=C6
    GRadio_980.place(x=380,y=170,width=100,height=30)

    GRadio_296=Radiobutton(root)
    GRadio_296["font"] = ('微軟正黑體',12)
    GRadio_296["justify"] = "left"
    GRadio_296["text"] = "NE無使用"
    GRadio_296['variable']=varNE
    GRadio_296['value']='0'
    GRadio_296['command']=C7
    GRadio_296.place(x=380,y=230,width=100,height=30)

    GRadio_262=Radiobutton(root)
    GRadio_262["font"] = ('微軟正黑體',12)
    GRadio_262["justify"] = "left"
    GRadio_262["text"] = "NP無使用"
    GRadio_262['variable']=varNP
    GRadio_262['value']='0'
    GRadio_262['command']=C8
    GRadio_262.place(x=380,y=290,width=100,height=30)
Right()
#顯示文字
def Txt():
    GLabel_163=Label(root)
    GLabel_163["font"] = ('微軟正黑體',10)
    GLabel_163["justify"] = "left"
    GLabel_163["text"] = "今日日期: "+datetime.strftime(datetime.today(),'%Y/%m/%d')
    GLabel_163.place(x=337,y=5,width=150,height=25)

    GLabel_94=Label(root)
    GLabel_94["font"] = ('微軟正黑體',14)
    GLabel_94["justify"] = "center"
    GLabel_94['anchor']='c'
    GLabel_94["text"] = "請選擇想要匯入的檔案，檔案打開時必須位於目標標籤頁"
    GLabel_94.place(x=0,y=60,width=500,height=30)
Txt()
root.mainloop()
