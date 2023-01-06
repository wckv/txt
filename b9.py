from os import path
from tkinter import Button,Checkbutton,Entry,Label,StringVar,Tk,messagebox,ttk
from datetime import datetime, timedelta
from openpyxl import load_workbook,styles
from send2trash import send2trash
root=Tk()

dM=2568721
dP='GM7173'
dR='10049M'

LotM,LotP,LotR,t=StringVar(),StringVar(),StringVar(),StringVar()
def Run():
    
    today=datetime.today()
    yesterday=today+timedelta(days=-1)
    today_xlsx='NB_COVID_'+datetime.strftime(today,'%m%d')+'.xlsx'
    yesterday_xlsx='NB_COVID_'+datetime.strftime(yesterday,'%m%d')+'.xlsx'

    if path.exists('./COVID/')==False:
        messagebox.showerror(title='錯誤',message='找不到COVID資料夾')
        return
    if path.exists('./公費克流感使用清單.txt')==False:
        messagebox.showerror(title='錯誤',message='找不到公費克流感使用清單.txt')
        return
    if path.exists('./COVID/'+yesterday_xlsx)==False:
        messagebox.showerror(title='錯誤',message='找不到昨日的Excel檔')
        return
    p['value']=5
    if path.isfile('./COVID/'+today_xlsx):
        send2trash('./COVID/'+today_xlsx)
        messagebox.showinfo(title='發現已執行過',message='原檔丟進資源回收桶後重新執行')
    try:
        with open('公費克流感使用清單.txt','r',encoding='big5',errors='ignore') as input:
            input=input.read()[101:]
            n=int(input.count('#')/19)
            input=input.split('#')
            lst=list(zip(input[1:19*n:19],input[4:19*n:19],input[8:19*n:19]))
        lstM=[x for x in lst if x[2]=='P4A151M']
        lstP=[x for x in lst if x[2]=='P4A152M']
        lstR=[x for x in lst if x[2]=='P4A135P']
        n=[len(lstM),len(lstP),len(lstR)]
        p['value']=10
        
        wb=load_workbook('./COVID/'+yesterday_xlsx)
        ws0=wb.worksheets[0]
        ws0.title='NB'+datetime.strftime(today,'%m%d')+'('+datetime.strftime(yesterday,'%m%d')+'庫存)'
        ws1=wb.worksheets[1]
        ws1.title='NB(P4A135P)'+datetime.strftime(today,'%m%d')+'('+datetime.strftime(yesterday,'%m%d')+'庫存)'
        B=[]
        for row in ws0['B1':'B'+str(ws0.max_row+1)]:
            for cell in row:
                if cell.value==None:
                    B+=[cell.row]
        BM=min(B)
        for i in range(len(B)-1):
            if (B[i+1]-B[i])!=1:
                BP=B[i+1]
        B=[]
        for row in ws1['B1':'B'+str(ws1.max_row+1)]:
            for cell in row:
                if cell.value==None:
                    B+=[cell.row]
        BR=min(B)

        ws0['L'+str(BM-1)].value=0 if ws0['L'+str(BM-1)].value==None else ws0['L'+str(BM-1)].value
        ws0['L'+str(BP-1)].value=0 if ws0['L'+str(BP-1)].value==None else ws0['L'+str(BP-1)].value
        ws1['L'+str(BR-1)].value=0 if ws1['L'+str(BR-1)].value==None else ws1['L'+str(BR-1)].value

        p['value']=30
        if n[0]==0:
            ws0.insert_rows(BM)
            ws0['A'+str(BM)]=1
            ws0['B'+str(BM)]=datetime.strftime(today,'%Y/%m/%d')
            ws0['C'+str(BM)]=datetime.strftime(yesterday,'%Y/%m/%d')
            ws0['D'+str(BM)]='高雄長庚紀念醫院'
            ws0['E'+str(BM)]='P4A151M'
            ws0['F'+str(BM)]='Molnupiravir'
            ws0['G'+str(BM)]=LotM.get()
            ws0['J'+str(BM)]='NB'
            ws0['K'+str(BM)]='無使用'
            ws0['L'+str(BM)]=ws0['L'+str(BM-1)].value
        else:
            for x in range(n[0]):
                ws0.insert_rows(BM+x)
                ws0['A'+str(BM+x)]=x+1
                ws0['B'+str(BM+x)]=datetime.strftime(today,'%Y/%m/%d')
                ws0['C'+str(BM+x)]=datetime.strftime(yesterday,'%Y/%m/%d')
                ws0['D'+str(BM+x)]='高雄長庚紀念醫院'
                ws0['E'+str(BM+x)]='P4A151M'
                ws0['F'+str(BM+x)]='Molnupiravir'
                ws0['G'+str(BM+x)]=LotM.get()
                ws0['H'+str(BM+x)]=lstM[x][0]
                ws0['I'+str(BM+x)]=lstM[x][1]
                ws0['J'+str(BM+x)]='NB'
            ws0['K'+str(BM+n[0]-1)]=n[0]
            ws0['L'+str(BM+n[0]-1)]=ws0['L'+str(BM-1)].value-n[0]

        BP+=1 if n[0]==0 else n[0]
        p['value']=50
        if n[1]==0:
            ws0.insert_rows(BP)
            ws0['A'+str(BP)]=1
            ws0['B'+str(BP)]=datetime.strftime(today,'%Y/%m/%d')
            ws0['C'+str(BP)]=datetime.strftime(yesterday,'%Y/%m/%d')
            ws0['D'+str(BP)]='高雄長庚紀念醫院'
            ws0['E'+str(BP)]='P4A152M'
            ws0['F'+str(BP)]='Paxlovid'
            ws0['G'+str(BP)]=LotP.get()
            ws0['J'+str(BP)]='NB'
            ws0['K'+str(BP)]='無使用'
            ws0['L'+str(BP)]=ws0['L'+str(BP-1)].value
        else:   
            for x in range(n[1]):
                ws0.insert_rows(BP+x)
                ws0['A'+str(BP+x)]=x+1
                ws0['B'+str(BP+x)]=datetime.strftime(today,'%Y/%m/%d')
                ws0['C'+str(BP+x)]=datetime.strftime(yesterday,'%Y/%m/%d')
                ws0['D'+str(BP+x)]='高雄長庚紀念醫院'
                ws0['E'+str(BP+x)]='P4A152M'
                ws0['F'+str(BP+x)]='Paxlovid'
                ws0['G'+str(BP+x)]=LotP.get()
                ws0['H'+str(BP+x)]=lstP[x][0]
                ws0['I'+str(BP+x)]=lstP[x][1]
                ws0['J'+str(BP+x)]='NB'
            ws0['K'+str(BP+n[1]-1)]=n[1]
            ws0['L'+str(BP+n[1]-1)]=ws0['L'+str(BP-1)].value-n[1]

        if n[2]==0:
            ws1.insert_rows(BR)
            ws1['A'+str(BR)]=1
            ws1['B'+str(BR)]=datetime.strftime(today,'%Y/%m/%d')
            ws1['C'+str(BR)]=datetime.strftime(yesterday,'%Y/%m/%d')
            ws1['D'+str(BR)]='高雄長庚紀念醫院'
            ws1['E'+str(BR)]='P4A135P'
            ws1['F'+str(BR)]='Remdesivir100mg/vial'
            ws1['G'+str(BR)]=LotR.get()
            ws1['K'+str(BR)]='NB'
            ws1['L'+str(BR)]='無使用'
            ws1['M'+str(BR)]=ws1['M'+str(BR-1)].value
        else:   
            for x in range(n[2]):
                ws1.insert_rows(BR+x)
                ws1['A'+str(BR+x)]=x+1
                ws1['B'+str(BR+x)]=datetime.strftime(today,'%Y/%m/%d')
                ws1['C'+str(BR+x)]=datetime.strftime(yesterday,'%Y/%m/%d')
                ws1['D'+str(BR+x)]='高雄長庚紀念醫院'
                ws1['E'+str(BR+x)]='P4A135P'
                ws1['F'+str(BR+x)]='Remdesivir100mg/vial'
                ws1['G'+str(BR+x)]=LotR.get()
                ws1['I'+str(BR+x)]=lstR[x][0]
                ws1['J'+str(BR+x)]=lstR[x][1]
                ws1['K'+str(BR+x)]='NB'
            ws1['L'+str(BR+n[2]-1)]=n[2]*6
            ws1['M'+str(BR+n[2]-1)]=ws1['M'+str(BR-1)].value-n[2]*6
        p['value']=70

        f=styles.Font(name='標楷體',size=12)
        ali=styles.Alignment(horizontal='right')
        line=styles.Border(left=styles.Side(style='thin'),right=styles.Side(style='thin'),top=styles.Side(style='thin'),
        bottom=styles.Side(style='thin'))

        for x in range(13):
            n[0]=1 if n[0]==0 else n[0]
            n[1]=1 if n[1]==0 else n[1]
            for y in range(n[0]):
                ws0[chr(65+x)+str(BM+y)].font=f
                ws0[chr(65+x)+str(BM+y)].border=line
                ws0[chr(65+x)+str(BM+y)].alignment=ali
            for y in range(n[1]):
                ws0[chr(65+x)+str(BP+y)].font=f
                ws0[chr(65+x)+str(BP+y)].border=line
                ws0[chr(65+x)+str(BP+y)].alignment=ali

        for x in range(14):
            n[2]=1 if n[2]==0 else n[2]
            for y in range(n[2]):
                ws1[chr(65+x)+str(BR+y)].font=f
                ws1[chr(65+x)+str(BR+y)].border=line
                ws1[chr(65+x)+str(BR+y)].alignment=ali
        p['value']= 90 #存檔
        wb.save('./COVID/'+today_xlsx)

        if check.get()=='1':
            send2trash('./公費克流感使用清單.txt')
        p['value']=100  #刪除檔案
        messagebox.showinfo(title='寫入成功',message='執行完畢')
        p['value']=0
    except:
        messagebox.showerror(title='發生錯誤',message='發生奇怪的錯誤')
        p['value']=0

def Info():
    messagebox.showinfo(title='說明',
    message='''
2023.01.05 P預設批號GM7173

4.2.2
1. 修正小錯誤(無使用時歪掉)(沒有值不能減)

4.1
1. 可手動輸入批號

4.0 新功能
1. 有視窗跟進度條。
2. 改成用插入資料的就不會有內容爆表的問題。
3. 如果有匯出瑞德希韋的話現在也會自動填入了。
[注意]4. 因為瑞德希韋匯出沒有病歷號碼，所以要手動輸入。

使用說明
1. HIS填的批號不會匯出所以隨便寫就好。
2. 如果是不須匯出的資料，批號直接空白就好了。
4. 將檔案存在桌面，檔名不用改。
5. 會讀取昨日日期檔案，複製成今日檔案後存檔。
6. 日期預設為電腦時間。
7. 匯出檔刪除是丟進垃圾桶，還可從垃圾桶中救回。
8. 如果輸出檔案已經存在，會先丟進垃圾桶。
''')

def time(): 
    now='[現在時間] ' + datetime.strftime(datetime.now(),'%Y/%m/%d %H:%M:%S')
    t.set(now)
    root.after(1000,time)

def GUI():
    global check
    check=StringVar()
    root.title('按我4.2')
    width=300
    height=200
    left=int((root.winfo_screenwidth()-width)/2)
    top=int((root.winfo_screenheight()-height)/2)
    root.geometry(f'{width}x{height}+{left}+{top}')
    root.resizable(False,False)

    global p
    p=ttk.Progressbar(root,length=280,orient='horizontal')
    p.place(x=10,y=170)
    
    txt1=Label(root,textvariable=t,font=('微軟正黑體',12))
    txt1.place(x=30,y=10)
    txt2=Label(root,text='M                   P                   R',font=('微軟正黑體',12))
    txt2.place(x=15,y=50)

    ent1=Entry(root,width=9,textvariable=LotM,font=('微軟正黑體',9))
    ent1.insert(0,dM)
    ent1.place(x=35,y=53)
    ent2=Entry(root,width=9,textvariable=LotP,font=('微軟正黑體',9))
    ent2.insert(0,dP)
    ent2.place(x=123,y=53)
    ent3=Entry(root,width=10,textvariable=LotR,font=('微軟正黑體',9))
    ent3.insert(0,dR)
    ent3.place(x=210,y=53)

    btn1=Button(root,width=5,text='說明',font=('微軟正黑體',12),command=Info)
    btn1.place(x=15,y=125)
    btn2=Button(root,width=5,text='執行',font=('微軟正黑體',12),command=Run)
    btn2.place(x=150,y=125)
    btn3=Button(root,width=5,text='離開',font=('微軟正黑體',12),command=lambda:root.destroy())
    btn3.place(x=225,y=125)

    de=Checkbutton(root,variable=check,font=('微軟正黑體',12),text='完成後刪除"公費克流感使用清單"')
    de.select()
    de.place(x=15,y=85)
time(),GUI(),root.mainloop()
